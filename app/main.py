from __future__ import annotations

import csv
import hashlib
import hmac
import io
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import quote_plus
from zoneinfo import ZoneInfo

from fastapi import FastAPI, Form, HTTPException, Request, Response, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, Field, ValidationError, field_validator
from sqlalchemy import ForeignKey, Index, String, Text, create_engine, func, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import DeclarativeBase, Mapped, Session, mapped_column, relationship, sessionmaker

BASE_DIR = Path(__file__).resolve().parent
PROJECT_DIR = BASE_DIR.parent
DEFAULT_DB_PATH = PROJECT_DIR / "data" / "playlist_event.db"
DEFAULT_SQLITE_URL = f"sqlite:///{DEFAULT_DB_PATH}"

DEFAULT_ADMIN_PASSWORD = "change-me-admin"
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", DEFAULT_ADMIN_PASSWORD)
SECRET_KEY = os.getenv("SECRET_KEY", "playlist-event-dev-secret")
ADMIN_COOKIE_NAME = "playlist_admin_session"
ADMIN_COOKIE_MAX_AGE = 60 * 60 * 12
SEOUL_TZ = ZoneInfo("Asia/Seoul")


class Base(DeclarativeBase):
    pass


class Entry(Base):
    __tablename__ = "entries"

    id: Mapped[int] = mapped_column(primary_key=True, autoincrement=True)
    name: Mapped[str] = mapped_column(String(40), nullable=False)
    company: Mapped[str] = mapped_column(String(60), nullable=False)
    department: Mapped[str] = mapped_column(String(60), nullable=False)
    song_title: Mapped[str] = mapped_column(String(120), nullable=False)
    artist_name: Mapped[str] = mapped_column(String(120), nullable=False)
    reason: Mapped[str] = mapped_column(Text, nullable=False)
    created_at: Mapped[str] = mapped_column(String(40), nullable=False)

    likes: Mapped[list["Like"]] = relationship(
        back_populates="entry",
        cascade="all, delete-orphan",
    )


class Like(Base):
    __tablename__ = "likes"
    __table_args__ = (
        Index("idx_likes_entry_id_token", "entry_id", "client_token", unique=True),
    )

    id: Mapped[int] = mapped_column(primary_key=True, autoincrement=True)
    entry_id: Mapped[int] = mapped_column(ForeignKey("entries.id", ondelete="CASCADE"), nullable=False)
    client_token: Mapped[str] = mapped_column(String(120), nullable=False)
    created_at: Mapped[str] = mapped_column(String(40), nullable=False)

    entry: Mapped[Entry] = relationship(back_populates="likes")


Index("idx_entries_created_at", Entry.created_at.desc())


def normalize_database_url() -> str:
    database_url = os.getenv("DATABASE_URL", "").strip()
    if database_url:
        if database_url.startswith("postgres://"):
            return database_url.replace("postgres://", "postgresql://", 1)
        return database_url

    db_path_env = os.getenv("DB_PATH", str(DEFAULT_DB_PATH))
    db_path = Path(db_path_env)
    db_path.parent.mkdir(parents=True, exist_ok=True)
    return f"sqlite:///{db_path}"


DATABASE_URL = normalize_database_url()
IS_SQLITE = DATABASE_URL.startswith("sqlite")
ENGINE_KWARGS: dict[str, Any] = {"future": True, "pool_pre_ping": True}
if IS_SQLITE:
    ENGINE_KWARGS["connect_args"] = {"check_same_thread": False}

engine = create_engine(DATABASE_URL, **ENGINE_KWARGS)
SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False, expire_on_commit=False)

app = FastAPI(title="Only One Playlist Event")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))


class EntryCreate(BaseModel):
    name: str = Field(min_length=1, max_length=40)
    company: str = Field(min_length=1, max_length=60)
    department: str = Field(min_length=1, max_length=60)
    song_title: str = Field(min_length=1, max_length=120, alias="songTitle")
    artist_name: str = Field(min_length=1, max_length=120, alias="artistName")
    reason: str = Field(min_length=10, max_length=500)

    @field_validator("name", "company", "department", "song_title", "artist_name", "reason")
    @classmethod
    def strip_and_validate(cls, value: str) -> str:
        cleaned = " ".join(value.split())
        if not cleaned:
            raise ValueError("빈 값은 허용되지 않습니다.")
        return cleaned


class LikePayload(BaseModel):
    client_token: str = Field(min_length=8, max_length=120, alias="clientToken")

    @field_validator("client_token")
    @classmethod
    def clean_token(cls, value: str) -> str:
        cleaned = value.strip()
        if not cleaned:
            raise ValueError("유효하지 않은 토큰입니다.")
        return cleaned


EXPORT_HEADERS = [
    "ID",
    "이름",
    "소속회사",
    "부서명",
    "추천곡명",
    "가수명",
    "추천 이유",
    "좋아요 수",
    "등록일시(KST)",
    "유튜브 검색 링크",
]


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()



def init_db() -> None:
    Base.metadata.create_all(bind=engine)



def build_youtube_search_url(song_title: str, artist_name: str) -> str:
    query = quote_plus(f"{song_title} {artist_name}")
    return f"https://www.youtube.com/results?search_query={query}"



def created_at_to_kst(created_at: str) -> str:
    try:
        dt = datetime.fromisoformat(created_at)
    except ValueError:
        return created_at

    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(SEOUL_TZ).strftime("%Y-%m-%d %H:%M:%S KST")



def row_mapping_to_dict(row: Any) -> dict[str, Any]:
    data = row._mapping if hasattr(row, "_mapping") else row
    song_title = data["song_title"]
    artist_name = data["artist_name"]
    created_at = data["created_at"]
    return {
        "id": data["id"],
        "name": data["name"],
        "company": data["company"],
        "department": data["department"],
        "songTitle": song_title,
        "artistName": artist_name,
        "reason": data["reason"],
        "createdAt": created_at,
        "createdAtKst": created_at_to_kst(created_at),
        "likes": int(data["likes_count"] or 0),
        "youtubeSearchUrl": build_youtube_search_url(song_title, artist_name),
    }



def entry_select(order_for_admin: bool = False):
    likes_count = func.count(Like.id).label("likes_count")
    stmt = (
        select(
            Entry.id,
            Entry.name,
            Entry.company,
            Entry.department,
            Entry.song_title,
            Entry.artist_name,
            Entry.reason,
            Entry.created_at,
            likes_count,
        )
        .select_from(Entry)
        .outerjoin(Like, Like.entry_id == Entry.id)
        .group_by(
            Entry.id,
            Entry.name,
            Entry.company,
            Entry.department,
            Entry.song_title,
            Entry.artist_name,
            Entry.reason,
            Entry.created_at,
        )
    )
    if order_for_admin:
        return stmt.order_by(Entry.created_at.desc())
    return stmt.order_by(likes_count.desc(), Entry.created_at.desc())



def fetch_entries() -> list[dict[str, Any]]:
    with SessionLocal() as session:
        rows = session.execute(entry_select(order_for_admin=False)).all()
    return [row_mapping_to_dict(row) for row in rows]



def fetch_admin_entries() -> list[dict[str, Any]]:
    with SessionLocal() as session:
        rows = session.execute(entry_select(order_for_admin=True)).all()
    return [row_mapping_to_dict(row) for row in rows]




def fetch_entry_by_id(entry_id: int) -> dict[str, Any] | None:
    with SessionLocal() as session:
        row = session.execute(entry_select(order_for_admin=True).where(Entry.id == entry_id)).first()
    if not row:
        return None
    return row_mapping_to_dict(row)



def fetch_stats() -> dict[str, int]:
    with SessionLocal() as session:
        total_entries = session.scalar(select(func.count()).select_from(Entry)) or 0
        total_likes = session.scalar(select(func.count()).select_from(Like)) or 0
        total_companies = session.scalar(select(func.count(func.distinct(Entry.company))).select_from(Entry)) or 0
    return {
        "totalEntries": int(total_entries),
        "totalLikes": int(total_likes),
        "totalCompanies": int(total_companies),
    }



def make_admin_cookie_value() -> str:
    payload = "playlist-admin-auth"
    digest = hmac.new(
        (SECRET_KEY + ADMIN_PASSWORD).encode("utf-8"),
        payload.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()
    return f"{payload}.{digest}"



def is_admin_authenticated(request: Request) -> bool:
    cookie_value = request.cookies.get(ADMIN_COOKIE_NAME)
    if not cookie_value:
        return False
    return hmac.compare_digest(cookie_value, make_admin_cookie_value())



def require_admin(request: Request) -> None:
    if not is_admin_authenticated(request):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="관리자 인증이 필요합니다.")



def set_admin_cookie(response: Response) -> None:
    response.set_cookie(
        key=ADMIN_COOKIE_NAME,
        value=make_admin_cookie_value(),
        httponly=True,
        samesite="lax",
        secure=False,
        max_age=ADMIN_COOKIE_MAX_AGE,
        path="/",
    )



def clear_admin_cookie(response: Response) -> None:
    response.delete_cookie(key=ADMIN_COOKIE_NAME, path="/")



def build_export_rows(items: list[dict[str, Any]]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for item in items:
        rows.append(
            [
                item["id"],
                item["name"],
                item["company"],
                item["department"],
                item["songTitle"],
                item["artistName"],
                item["reason"],
                item["likes"],
                item["createdAtKst"],
                item["youtubeSearchUrl"],
            ]
        )
    return rows



def build_admin_template_context(
    request: Request,
    *,
    authenticated: bool,
    error_message: str | None = None,
) -> dict[str, Any]:
    items = fetch_admin_entries() if authenticated else []
    stats = fetch_stats()
    return {
        "page_title": "Playlist Event Admin",
        "authenticated": authenticated,
        "items": items,
        "stats": stats,
        "error_message": error_message,
        "default_password_in_use": ADMIN_PASSWORD == DEFAULT_ADMIN_PASSWORD,
        "request": request,
    }


@app.on_event("startup")
def startup() -> None:
    init_db()


@app.get("/", response_class=HTMLResponse)
def index(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(
        request=request,
        name="index.html",
        context={
            "page_title": "Only One Playlist",
            "event_title": "MY FAVORITE PLAYLIST",
            "event_label": "only one",
        },
    )


@app.get("/admin", response_class=HTMLResponse)
def admin_page(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(
        name="admin.html",
        context=build_admin_template_context(
            request,
            authenticated=is_admin_authenticated(request),
        ),
        request=request,
    )


@app.post("/admin/login", response_class=HTMLResponse)
def admin_login(request: Request, password: str = Form(...)) -> Response:
    if password != ADMIN_PASSWORD:
        return templates.TemplateResponse(
            name="admin.html",
            context=build_admin_template_context(
                request,
                authenticated=False,
                error_message="비밀번호가 올바르지 않습니다.",
            ),
            request=request,
            status_code=status.HTTP_401_UNAUTHORIZED,
        )

    response = RedirectResponse(url="/admin", status_code=status.HTTP_303_SEE_OTHER)
    set_admin_cookie(response)
    return response


@app.post("/admin/logout")
def admin_logout() -> Response:
    response = RedirectResponse(url="/admin", status_code=status.HTTP_303_SEE_OTHER)
    clear_admin_cookie(response)
    return response



@app.post("/admin/entries/{entry_id}/delete")
def admin_delete_entry(entry_id: int, request: Request) -> Response:
    require_admin(request)
    with SessionLocal() as session:
        entry = session.get(Entry, entry_id)
        if entry is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="항목을 찾을 수 없습니다.")
        session.delete(entry)
        session.commit()
    return RedirectResponse(url="/admin", status_code=status.HTTP_303_SEE_OTHER)



@app.get("/admin/entries/{entry_id}/edit", response_class=HTMLResponse)
def admin_edit_entry_page(entry_id: int, request: Request) -> HTMLResponse:
    require_admin(request)
    item = fetch_entry_by_id(entry_id)
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="항목을 찾을 수 없습니다.")
    return templates.TemplateResponse(
        request=request,
        name="admin_edit.html",
        context={
            "page_title": "Playlist Entry Edit",
            "item": item,
            "error_message": None,
        },
    )



@app.post("/admin/entries/{entry_id}/edit")
def admin_edit_entry_submit(
    entry_id: int,
    request: Request,
    name: str = Form(...),
    company: str = Form(...),
    department: str = Form(...),
    song_title: str = Form(...),
    artist_name: str = Form(...),
    reason: str = Form(...),
) -> Response:
    require_admin(request)

    try:
        payload = EntryCreate(
            name=name,
            company=company,
            department=department,
            songTitle=song_title,
            artistName=artist_name,
            reason=reason,
        )
    except ValidationError:
        item = fetch_entry_by_id(entry_id) or {
            "id": entry_id,
            "name": name,
            "company": company,
            "department": department,
            "songTitle": song_title,
            "artistName": artist_name,
            "reason": reason,
            "likes": 0,
            "createdAtKst": "",
            "youtubeSearchUrl": build_youtube_search_url(song_title, artist_name),
        }
        return templates.TemplateResponse(
            request=request,
            name="admin_edit.html",
            context={
                "page_title": "Playlist Entry Edit",
                "item": item,
                "error_message": "입력값을 다시 확인해 주세요. (추천 이유는 10자 이상, 500자 이하)",
            },
            status_code=status.HTTP_400_BAD_REQUEST,
        )

    with SessionLocal() as session:
        entry = session.get(Entry, entry_id)
        if entry is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="항목을 찾을 수 없습니다.")

        duplicate_stmt = (
            select(Entry.id)
            .where(func.lower(Entry.name) == payload.name.lower())
            .where(func.lower(Entry.company) == payload.company.lower())
            .where(func.lower(Entry.department) == payload.department.lower())
            .where(Entry.id != entry_id)
            .limit(1)
        )
        duplicate = session.scalar(duplicate_stmt)
        if duplicate is not None:
            item = fetch_entry_by_id(entry_id)
            if item is None:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="항목을 찾을 수 없습니다.")
            item["name"] = payload.name
            item["company"] = payload.company
            item["department"] = payload.department
            item["songTitle"] = payload.song_title
            item["artistName"] = payload.artist_name
            item["reason"] = payload.reason
            item["youtubeSearchUrl"] = build_youtube_search_url(payload.song_title, payload.artist_name)

            return templates.TemplateResponse(
                request=request,
                name="admin_edit.html",
                context={
                    "page_title": "Playlist Entry Edit",
                    "item": item,
                    "error_message": "이름/회사/부서 기준으로 이미 다른 추천곡이 등록되어 있습니다.",
                },
                status_code=status.HTTP_409_CONFLICT,
            )

        entry.name = payload.name
        entry.company = payload.company
        entry.department = payload.department
        entry.song_title = payload.song_title
        entry.artist_name = payload.artist_name
        entry.reason = payload.reason
        session.commit()

    return RedirectResponse(url="/admin", status_code=status.HTTP_303_SEE_OTHER)



@app.get("/admin/export.csv")
def export_csv(request: Request) -> Response:
    require_admin(request)
    items = fetch_admin_entries()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(EXPORT_HEADERS)
    writer.writerows(build_export_rows(items))
    filename = f"playlist-event-export-{datetime.now(SEOUL_TZ).strftime('%Y%m%d-%H%M%S')}.csv"
    content = "\ufeff" + output.getvalue()
    return Response(
        content=content,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/admin/export.xlsx")
def export_xlsx(request: Request) -> StreamingResponse:
    require_admin(request)
    items = fetch_admin_entries()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Playlist Entries"
    sheet.append(EXPORT_HEADERS)
    for row in build_export_rows(items):
        sheet.append(row)

    header_fill = PatternFill(fill_type="solid", fgColor="0F4478")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        "A": 8,
        "B": 14,
        "C": 18,
        "D": 18,
        "E": 24,
        "F": 22,
        "G": 46,
        "H": 12,
        "I": 22,
        "J": 50,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    sheet.freeze_panes = "A2"

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename = f"playlist-event-export-{datetime.now(SEOUL_TZ).strftime('%Y%m%d-%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/api/entries")
def list_entries() -> dict[str, Any]:
    return {
        "items": fetch_entries(),
        "stats": fetch_stats(),
    }


@app.post("/api/entries")
def create_entry(payload: EntryCreate) -> dict[str, Any]:
    created_at = utc_now_iso()
    with SessionLocal() as session:
        duplicate_stmt = (
            select(Entry.id)
            .where(func.lower(Entry.name) == payload.name.lower())
            .where(func.lower(Entry.company) == payload.company.lower())
            .where(func.lower(Entry.department) == payload.department.lower())
            .limit(1)
        )
        duplicate = session.scalar(duplicate_stmt)
        if duplicate is not None:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="이름/회사/부서 기준으로 이미 추천곡이 등록되어 있습니다. 1인 1곡만 참여할 수 있어요.",
            )

        entry = Entry(
            name=payload.name,
            company=payload.company,
            department=payload.department,
            song_title=payload.song_title,
            artist_name=payload.artist_name,
            reason=payload.reason,
            created_at=created_at,
        )
        session.add(entry)
        session.commit()
        session.refresh(entry)

    item = {
        "id": entry.id,
        "name": entry.name,
        "company": entry.company,
        "department": entry.department,
        "songTitle": entry.song_title,
        "artistName": entry.artist_name,
        "reason": entry.reason,
        "createdAt": entry.created_at,
        "createdAtKst": created_at_to_kst(entry.created_at),
        "likes": 0,
        "youtubeSearchUrl": build_youtube_search_url(entry.song_title, entry.artist_name),
    }
    return {"item": item, "stats": fetch_stats()}


@app.post("/api/entries/{entry_id}/like")
def like_entry(entry_id: int, payload: LikePayload) -> dict[str, Any]:
    already_liked = False
    with SessionLocal() as session:
        entry = session.get(Entry, entry_id)
        if entry is None:
            raise HTTPException(status_code=404, detail="항목을 찾을 수 없습니다.")

        like = Like(
            entry_id=entry_id,
            client_token=payload.client_token,
            created_at=utc_now_iso(),
        )
        session.add(like)
        try:
            session.commit()
        except IntegrityError:
            session.rollback()
            already_liked = True

        likes_count = session.scalar(select(func.count()).select_from(Like).where(Like.entry_id == entry_id)) or 0

    response: dict[str, Any] = {
        "entryId": entry_id,
        "likes": int(likes_count),
        "alreadyLiked": already_liked,
        "stats": fetch_stats(),
    }
    return response
