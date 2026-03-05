from __future__ import annotations

import csv
import hashlib
import hmac
import io
import os
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import quote_plus
from zoneinfo import ZoneInfo

from fastapi import FastAPI, Form, HTTPException, Request, Response, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, Field, field_validator

BASE_DIR = Path(__file__).resolve().parent
PROJECT_DIR = BASE_DIR.parent
DEFAULT_DB_PATH = PROJECT_DIR / "data" / "playlist_event.db"
DB_PATH = Path(os.getenv("DB_PATH", str(DEFAULT_DB_PATH)))

DEFAULT_ADMIN_PASSWORD = "change-me-admin"
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", DEFAULT_ADMIN_PASSWORD)
SECRET_KEY = os.getenv("SECRET_KEY", "playlist-event-dev-secret")
ADMIN_COOKIE_NAME = "playlist_admin_session"
ADMIN_COOKIE_MAX_AGE = 60 * 60 * 12
SEOUL_TZ = ZoneInfo("Asia/Seoul")

app = FastAPI(title="Only One Playlist Event")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))


class EntryCreate(BaseModel):
    name: str = Field(min_length=1, max_length=40)
    company: str = Field(min_length=1, max_length=60)
    department: str = Field(min_length=1, max_length=60)
    song_title: str = Field(min_length=1, max_length=120, alias="songTitle")
    artist_name: str = Field(min_length=1, max_length=120, alias="artistName")
    reason: str = Field(min_length=10, max_length=500)

    @field_validator("name", "company", "department", "song_title", "artist_name", "reason")
    @classmethod
    def strip_and_validate(cls, value: str) -> str:
        cleaned = " ".join(value.split())
        if not cleaned:
            raise ValueError("빈 값은 허용되지 않습니다.")
        return cleaned


class LikePayload(BaseModel):
    client_token: str = Field(min_length=8, max_length=120, alias="clientToken")

    @field_validator("client_token")
    @classmethod
    def clean_token(cls, value: str) -> str:
        cleaned = value.strip()
        if not cleaned:
            raise ValueError("유효하지 않은 토큰입니다.")
        return cleaned


EXPORT_HEADERS = [
    "ID",
    "이름",
    "소속회사",
    "부서명",
    "추천곡명",
    "가수명",
    "추천 이유",
    "좋아요 수",
    "등록일시(KST)",
    "유튜브 검색 링크",
]


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()



def get_connection() -> sqlite3.Connection:
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn



def init_db() -> None:
    with get_connection() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                company TEXT NOT NULL,
                department TEXT NOT NULL,
                song_title TEXT NOT NULL,
                artist_name TEXT NOT NULL,
                reason TEXT NOT NULL,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS likes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                entry_id INTEGER NOT NULL,
                client_token TEXT NOT NULL,
                created_at TEXT NOT NULL,
                FOREIGN KEY (entry_id) REFERENCES entries(id) ON DELETE CASCADE,
                UNIQUE(entry_id, client_token)
            );

            CREATE INDEX IF NOT EXISTS idx_entries_created_at ON entries(created_at DESC);
            CREATE INDEX IF NOT EXISTS idx_likes_entry_id ON likes(entry_id);
            """
        )



def build_youtube_search_url(song_title: str, artist_name: str) -> str:
    query = quote_plus(f"{song_title} {artist_name}")
    return f"https://www.youtube.com/results?search_query={query}"



def created_at_to_kst(created_at: str) -> str:
    try:
        dt = datetime.fromisoformat(created_at)
    except ValueError:
        return created_at

    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(SEOUL_TZ).strftime("%Y-%m-%d %H:%M:%S KST")



def entry_row_to_dict(row: sqlite3.Row) -> dict[str, Any]:
    song_title = row["song_title"]
    artist_name = row["artist_name"]
    created_at = row["created_at"]
    return {
        "id": row["id"],
        "name": row["name"],
        "company": row["company"],
        "department": row["department"],
        "songTitle": song_title,
        "artistName": artist_name,
        "reason": row["reason"],
        "createdAt": created_at,
        "createdAtKst": created_at_to_kst(created_at),
        "likes": row["likes_count"],
        "youtubeSearchUrl": build_youtube_search_url(song_title, artist_name),
    }



def fetch_entries() -> list[dict[str, Any]]:
    query = """
        SELECT
            e.id,
            e.name,
            e.company,
            e.department,
            e.song_title,
            e.artist_name,
            e.reason,
            e.created_at,
            COUNT(l.id) AS likes_count
        FROM entries e
        LEFT JOIN likes l ON l.entry_id = e.id
        GROUP BY e.id
        ORDER BY likes_count DESC, e.created_at DESC
    """
    with get_connection() as conn:
        rows = conn.execute(query).fetchall()
    return [entry_row_to_dict(row) for row in rows]



def fetch_admin_entries() -> list[dict[str, Any]]:
    query = """
        SELECT
            e.id,
            e.name,
            e.company,
            e.department,
            e.song_title,
            e.artist_name,
            e.reason,
            e.created_at,
            COUNT(l.id) AS likes_count
        FROM entries e
        LEFT JOIN likes l ON l.entry_id = e.id
        GROUP BY e.id
        ORDER BY e.created_at DESC
    """
    with get_connection() as conn:
        rows = conn.execute(query).fetchall()
    return [entry_row_to_dict(row) for row in rows]



def fetch_stats() -> dict[str, int]:
    with get_connection() as conn:
        total_entries = conn.execute("SELECT COUNT(*) FROM entries").fetchone()[0]
        total_likes = conn.execute("SELECT COUNT(*) FROM likes").fetchone()[0]
        total_companies = conn.execute("SELECT COUNT(DISTINCT company) FROM entries").fetchone()[0]
    return {
        "totalEntries": total_entries,
        "totalLikes": total_likes,
        "totalCompanies": total_companies,
    }



def make_admin_cookie_value() -> str:
    payload = "playlist-admin-auth"
    digest = hmac.new(
        (SECRET_KEY + ADMIN_PASSWORD).encode("utf-8"),
        payload.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()
    return f"{payload}.{digest}"



def is_admin_authenticated(request: Request) -> bool:
    cookie_value = request.cookies.get(ADMIN_COOKIE_NAME)
    if not cookie_value:
        return False
    return hmac.compare_digest(cookie_value, make_admin_cookie_value())



def require_admin(request: Request) -> None:
    if not is_admin_authenticated(request):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="관리자 인증이 필요합니다.")



def set_admin_cookie(response: Response) -> None:
    response.set_cookie(
        key=ADMIN_COOKIE_NAME,
        value=make_admin_cookie_value(),
        httponly=True,
        samesite="lax",
        secure=False,
        max_age=ADMIN_COOKIE_MAX_AGE,
        path="/",
    )



def clear_admin_cookie(response: Response) -> None:
    response.delete_cookie(key=ADMIN_COOKIE_NAME, path="/")



def build_export_rows(items: list[dict[str, Any]]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for item in items:
        rows.append(
            [
                item["id"],
                item["name"],
                item["company"],
                item["department"],
                item["songTitle"],
                item["artistName"],
                item["reason"],
                item["likes"],
                item["createdAtKst"],
                item["youtubeSearchUrl"],
            ]
        )
    return rows



def build_admin_template_context(
    request: Request,
    *,
    authenticated: bool,
    error_message: str | None = None,
) -> dict[str, Any]:
    items = fetch_admin_entries() if authenticated else []
    stats = fetch_stats()
    return {
        "page_title": "Playlist Event Admin",
        "authenticated": authenticated,
        "items": items,
        "stats": stats,
        "error_message": error_message,
        "default_password_in_use": ADMIN_PASSWORD == DEFAULT_ADMIN_PASSWORD,
        "request": request,
    }


@app.on_event("startup")
def startup() -> None:
    init_db()


@app.get("/", response_class=HTMLResponse)
def index(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(
        request=request,
        name="index.html",
        context={
            "page_title": "Only One Playlist",
            "event_title": "MY FAVORITE PLAYLIST",
            "event_label": "only one",
        },
    )


@app.get("/admin", response_class=HTMLResponse)
def admin_page(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(
        name="admin.html",
        context=build_admin_template_context(
            request,
            authenticated=is_admin_authenticated(request),
        ),
        request=request,
    )


@app.post("/admin/login", response_class=HTMLResponse)
def admin_login(request: Request, password: str = Form(...)) -> Response:
    if password != ADMIN_PASSWORD:
        return templates.TemplateResponse(
            name="admin.html",
            context=build_admin_template_context(
                request,
                authenticated=False,
                error_message="비밀번호가 올바르지 않습니다.",
            ),
            request=request,
            status_code=status.HTTP_401_UNAUTHORIZED,
        )

    response = RedirectResponse(url="/admin", status_code=status.HTTP_303_SEE_OTHER)
    set_admin_cookie(response)
    return response


@app.post("/admin/logout")
def admin_logout() -> Response:
    response = RedirectResponse(url="/admin", status_code=status.HTTP_303_SEE_OTHER)
    clear_admin_cookie(response)
    return response


@app.get("/admin/export.csv")
def export_csv(request: Request) -> Response:
    require_admin(request)
    items = fetch_admin_entries()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(EXPORT_HEADERS)
    writer.writerows(build_export_rows(items))
    filename = f"playlist-event-export-{datetime.now(SEOUL_TZ).strftime('%Y%m%d-%H%M%S')}.csv"
    content = "\ufeff" + output.getvalue()
    return Response(
        content=content,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/admin/export.xlsx")
def export_xlsx(request: Request) -> StreamingResponse:
    require_admin(request)
    items = fetch_admin_entries()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Playlist Entries"
    sheet.append(EXPORT_HEADERS)
    for row in build_export_rows(items):
        sheet.append(row)

    header_fill = PatternFill(fill_type="solid", fgColor="0F4478")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        "A": 8,
        "B": 14,
        "C": 18,
        "D": 18,
        "E": 24,
        "F": 22,
        "G": 46,
        "H": 12,
        "I": 22,
        "J": 50,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    sheet.freeze_panes = "A2"

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename = f"playlist-event-export-{datetime.now(SEOUL_TZ).strftime('%Y%m%d-%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/api/entries")
def list_entries() -> dict[str, Any]:
    return {
        "items": fetch_entries(),
        "stats": fetch_stats(),
    }


@app.post("/api/entries")
def create_entry(payload: EntryCreate) -> dict[str, Any]:
    created_at = utc_now_iso()
    with get_connection() as conn:
        duplicate = conn.execute(
            """
            SELECT id
            FROM entries
            WHERE LOWER(name) = LOWER(?)
              AND LOWER(company) = LOWER(?)
              AND LOWER(department) = LOWER(?)
            LIMIT 1
            """,
            (payload.name, payload.company, payload.department),
        ).fetchone()
        if duplicate is not None:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="이름/회사/부서 기준으로 이미 추천곡이 등록되어 있습니다. 1인 1곡만 참여할 수 있어요.",
            )

        cursor = conn.execute(
            """
            INSERT INTO entries (name, company, department, song_title, artist_name, reason, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                payload.name,
                payload.company,
                payload.department,
                payload.song_title,
                payload.artist_name,
                payload.reason,
                created_at,
            ),
        )
        entry_id = cursor.lastrowid
        row = conn.execute(
            """
            SELECT
                e.id,
                e.name,
                e.company,
                e.department,
                e.song_title,
                e.artist_name,
                e.reason,
                e.created_at,
                0 AS likes_count
            FROM entries e
            WHERE e.id = ?
            """,
            (entry_id,),
        ).fetchone()
    if row is None:
        raise HTTPException(status_code=500, detail="등록된 항목을 불러오지 못했습니다.")
    return {"item": entry_row_to_dict(row), "stats": fetch_stats()}


@app.post("/api/entries/{entry_id}/like")
def like_entry(entry_id: int, payload: LikePayload) -> dict[str, Any]:
    with get_connection() as conn:
        entry_exists = conn.execute("SELECT 1 FROM entries WHERE id = ?", (entry_id,)).fetchone()
        if entry_exists is None:
            raise HTTPException(status_code=404, detail="항목을 찾을 수 없습니다.")

        already_liked = False
        try:
            conn.execute(
                """
                INSERT INTO likes (entry_id, client_token, created_at)
                VALUES (?, ?, ?)
                """,
                (entry_id, payload.client_token, utc_now_iso()),
            )
        except sqlite3.IntegrityError:
            already_liked = True

        likes_count = conn.execute(
            "SELECT COUNT(*) FROM likes WHERE entry_id = ?",
            (entry_id,),
        ).fetchone()[0]

    response: dict[str, Any] = {
        "entryId": entry_id,
        "likes": likes_count,
        "alreadyLiked": already_liked,
        "stats": fetch_stats(),
    }
    return response
